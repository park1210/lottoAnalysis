from __future__ import annotations

import re
import warnings
from pathlib import Path

import pandas as pd
import requests
from bs4 import BeautifulSoup
from openpyxl import Workbook, load_workbook
from requests import HTTPError

from src.config import RAW_LOTTO_FILE


PYONY_ROUND_URL = "https://pyony.com/lotto/rounds/{round_no}/"
REQUEST_HEADERS = {
    "User-Agent": (
        "Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
        "AppleWebKit/537.36 (KHTML, like Gecko) "
        "Chrome/134.0.0.0 Safari/537.36"
    ),
    "Accept-Language": "ko-KR,ko;q=0.9,en-US;q=0.8,en;q=0.7",
}
RAW_LOTTO_COLUMNS = [
    "No",
    "Round",
    "Number1",
    "Number2",
    "Number3",
    "Number4",
    "Number5",
    "Number6",
    "Bonus",
    "Rank",
    "WinnerCount",
    "PrizeAmount",
]
RAW_LOTTO_DISPLAY_COLUMNS = [
    "No",
    "\ud68c\ucc28",
    "\ub2f9\ucca8\ubc88\ud6381",
    "\ub2f9\ucca8\ubc88\ud6382",
    "\ub2f9\ucca8\ubc88\ud6383",
    "\ub2f9\ucca8\ubc88\ud6384",
    "\ub2f9\ucca8\ubc88\ud6385",
    "\ub2f9\ucca8\ubc88\ud6386",
    "\ubcf4\ub108\uc2a4",
    "\uc21c\uc704",
    "\ub2f9\ucca8\uac8c\uc784\uc218",
    "1\uac8c\uc784\ub2f9 \ub2f9\ucca8\uae08\uc561",
]
LEGACY_KR_TO_CANONICAL = {
    "No": "No",
    "\ud68c\ucc28": "Round",
    "\ub2f9\ucca8\ubc88\ud638": "Number1",
    "\ub2f9\ucca8\ubc88\ud6381": "Number1",
    "\ub2f9\ucca8\ubc88\ud6382": "Number2",
    "\ub2f9\ucca8\ubc88\ud6383": "Number3",
    "\ub2f9\ucca8\ubc88\ud6384": "Number4",
    "\ub2f9\ucca8\ubc88\ud6385": "Number5",
    "\ub2f9\ucca8\ubc88\ud6386": "Number6",
    "Unnamed: 3": "Number2",
    "Unnamed: 4": "Number3",
    "Unnamed: 5": "Number4",
    "Unnamed: 6": "Number5",
    "Unnamed: 7": "Number6",
    "\ubcf4\ub108\uc2a4": "Bonus",
    "\uc21c\uc704": "Rank",
    "\ub2f9\ucca8\uac8c\uc784\uc218": "WinnerCount",
    "1\uac8c\uc784\ub2f9 \ub2f9\ucca8\uae08\uc561": "PrizeAmount",
}
RESULT_PENDING_TEXT = "\uacb0\uacfc\ub97c \uae30\ub2e4\ub9ac\uace0 \uc788\uc2b5\ub2c8\ub2e4."
ROUND_HEADING_SUFFIX = "\ud68c ("
DRAW_TEXT = "\ucd94\ucca8"
DISCLAIMER_PREFIX = "\ubcf8 \uc815\ubcf4\uc5d0\ub294 \uc624\ub958\uac00 \uc788\uc744 \uc218"
FIRST_RANK_TEXT = "1\ub4f1"
WINNER_SUFFIX = " \uba85"
PRIZE_SUFFIX = " \uc6d0"


def normalize_history_columns(df: pd.DataFrame) -> pd.DataFrame:
    if df.empty:
        return pd.DataFrame(columns=RAW_LOTTO_COLUMNS)

    normalized = df.copy().rename(columns=LEGACY_KR_TO_CANONICAL)
    if "Round" not in normalized.columns and len(normalized.columns) >= len(RAW_LOTTO_COLUMNS):
        normalized = normalized.iloc[:, : len(RAW_LOTTO_COLUMNS)].copy()
        normalized.columns = RAW_LOTTO_COLUMNS

    return normalized.reindex(columns=RAW_LOTTO_COLUMNS)


def load_existing_history(path: Path) -> pd.DataFrame:
    if not path.exists():
        return pd.DataFrame(columns=RAW_LOTTO_COLUMNS)

    return normalize_history_columns(pd.read_excel(path))


def get_last_local_round(df: pd.DataFrame) -> int:
    if df.empty:
        return 0

    rounds = pd.to_numeric(df["Round"], errors="coerce").dropna()
    if rounds.empty:
        return 0

    return int(rounds.max())


def _format_winner_count(value) -> str:
    numeric = pd.to_numeric(pd.Series([value]), errors="coerce").iloc[0]
    if pd.isna(numeric):
        return ""
    return f"{int(numeric):,}{WINNER_SUFFIX}"


def _format_prize_amount(value) -> str:
    numeric = pd.to_numeric(pd.Series([value]), errors="coerce").iloc[0]
    if pd.isna(numeric):
        return ""
    return f"{int(numeric):,}{PRIZE_SUFFIX}"


def _parse_excel_int(value) -> int | None:
    if value is None:
        return None
    text = str(value).replace(",", "").strip()
    if not text:
        return None
    try:
        return int(float(text))
    except ValueError:
        return None


def ensure_flat_headers(path: Path):
    if path.exists():
        workbook = load_workbook(path)
        worksheet = workbook.active
    else:
        path.parent.mkdir(parents=True, exist_ok=True)
        workbook = Workbook()
        worksheet = workbook.active

    for merged_range in list(worksheet.merged_cells.ranges):
        if merged_range.min_row == 1 and merged_range.max_row == 1:
            worksheet.unmerge_cells(str(merged_range))

    for idx, column_name in enumerate(RAW_LOTTO_DISPLAY_COLUMNS, start=1):
        worksheet.cell(row=1, column=idx, value=column_name)

    return workbook, worksheet


def get_workbook_state(path: Path) -> tuple[int, int]:
    workbook, worksheet = ensure_flat_headers(path)
    max_round = 0
    max_no = 0

    for row in worksheet.iter_rows(min_row=2, max_col=2, values_only=True):
        row_no, round_value = row
        parsed_no = _parse_excel_int(row_no)
        parsed_round = _parse_excel_int(round_value)
        if parsed_no is not None:
            max_no = max(max_no, parsed_no)
        if parsed_round is not None:
            max_round = max(max_round, parsed_round)

    workbook.save(path)
    workbook.close()
    return max_round, max_no


def append_history_rows(output_path: Path, records: list[list[int | str]]) -> Path:
    workbook, worksheet = ensure_flat_headers(output_path)

    last_no = 0
    for row_no, _ in worksheet.iter_rows(min_row=2, max_col=2, values_only=True):
        parsed_no = _parse_excel_int(row_no)
        if parsed_no is not None:
            last_no = max(last_no, parsed_no)

    next_no = last_no + 1
    for record in records:
        round_no = record[1]
        row_values = [
            next_no,
            round_no,
            record[2],
            record[3],
            record[4],
            record[5],
            record[6],
            record[7],
            record[8],
            FIRST_RANK_TEXT,
            _format_winner_count(record[10]),
            _format_prize_amount(record[11]),
        ]
        worksheet.append(row_values)
        next_no += 1

    workbook.save(output_path)
    workbook.close()
    return output_path


def fetch_round_page(round_no: int, timeout: int = 10) -> BeautifulSoup:
    response = requests.get(
        PYONY_ROUND_URL.format(round_no=round_no),
        headers=REQUEST_HEADERS,
        timeout=timeout,
    )
    try:
        response.raise_for_status()
    except HTTPError as exc:
        if response.status_code in {404, 500}:
            raise ValueError(f"Round {round_no} is not available") from exc
        raise
    return BeautifulSoup(response.text, "lxml")


def _extract_first_prize_values(text_tokens: list[str]) -> tuple[int, int]:
    joined = "\n".join(text_tokens)
    match = re.search(r"1\ub4f1\s+([\d,]+)\s+([\d,]+)\s*\uc6d0", joined)
    if not match:
        raise RuntimeError("Could not parse first prize data from the HTML result page")

    winner_count = int(match.group(1).replace(",", ""))
    prize_amount = int(match.group(2).replace(",", ""))
    return winner_count, prize_amount


def parse_round_record_from_html(round_no: int, soup: BeautifulSoup) -> list[int | str]:
    text_tokens = list(soup.stripped_strings)
    if any(token == RESULT_PENDING_TEXT for token in text_tokens):
        raise ValueError(f"Round {round_no} is not available yet")

    round_heading_index = next(
        (
            idx
            for idx, token in enumerate(text_tokens)
            if token.startswith(f"{round_no}{ROUND_HEADING_SUFFIX}") and DRAW_TEXT in token
        ),
        None,
    )
    if round_heading_index is None:
        raise RuntimeError(f"Could not locate the round heading for round {round_no}")

    number_tokens: list[int] = []
    for token in text_tokens[round_heading_index + 1 :]:
        if token.startswith(DISCLAIMER_PREFIX):
            break
        if token.isdigit():
            number_tokens.append(int(token))
        if len(number_tokens) == 7:
            break

    if len(number_tokens) != 7:
        raise RuntimeError(f"Could not parse the winning numbers for round {round_no}")

    winner_count, prize_amount = _extract_first_prize_values(text_tokens)
    return [
        round_no,
        round_no,
        number_tokens[0],
        number_tokens[1],
        number_tokens[2],
        number_tokens[3],
        number_tokens[4],
        number_tokens[5],
        number_tokens[6],
        FIRST_RANK_TEXT,
        winner_count,
        prize_amount,
    ]


def fetch_round_record(round_no: int) -> list[int | str]:
    soup = fetch_round_page(round_no)
    return parse_round_record_from_html(round_no, soup)


def is_round_available(round_no: int) -> bool:
    try:
        fetch_round_record(round_no)
        return True
    except ValueError:
        return False


def get_latest_available_round(start_round: int = 1) -> int:
    if start_round < 1:
        start_round = 1

    if not is_round_available(start_round):
        raise RuntimeError("Could not find any available lotto rounds")

    low = start_round
    high = start_round

    while is_round_available(high):
        low = high
        high *= 2

    while low + 1 < high:
        mid = (low + high) // 2
        if is_round_available(mid):
            low = mid
        else:
            high = mid

    return low


def fetch_missing_history(start_round: int, end_round: int) -> pd.DataFrame:
    records = [fetch_round_record(round_no) for round_no in range(start_round, end_round + 1)]
    return pd.DataFrame(records, columns=RAW_LOTTO_COLUMNS)


def fetch_incremental_history(start_round: int) -> pd.DataFrame:
    records: list[list[int | str]] = []
    round_no = max(1, start_round)
    print(f"Checking for new rounds starting from {round_no}...")

    while True:
        try:
            print(f"Fetching round {round_no}...")
            records.append(fetch_round_record(round_no))
            round_no += 1
        except ValueError:
            print(f"Round {round_no} is not available. Stopping incremental sync.")
            break

    return pd.DataFrame(records, columns=RAW_LOTTO_COLUMNS)


def sync_lotto_history_html(output_path: Path = RAW_LOTTO_FILE) -> Path:
    local_round, last_no = get_workbook_state(output_path)
    print(f"Last stored round: {local_round}")
    incremental_df = fetch_incremental_history(local_round + 1)

    if incremental_df.empty:
        print("No workbook updates were needed.")
        return output_path

    print(f"Appending {len(incremental_df)} new rows after No {last_no} into {output_path}...")
    return append_history_rows(output_path, incremental_df.values.tolist())


def ensure_lotto_history_available(output_path: Path = RAW_LOTTO_FILE) -> Path:
    try:
        return sync_lotto_history_html(output_path=output_path)
    except Exception as exc:
        if output_path.exists():
            warnings.warn(
                f"Could not sync lotto history from the remote HTML source. "
                f"Using the existing workbook instead. ({exc})",
                stacklevel=2,
            )
            return output_path
        raise
